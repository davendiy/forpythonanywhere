#!/usr/bin/env python3
# -*-encoding: utf-8-*-

# created: 04.10.2019
# by David Zashkolny
# 3 course, comp math
# Taras Shevchenko National University of Kyiv
# email: davendiy@gmail.com

import concurrent.futures
from tkinter import *
from tkinter import ttk
import openpyxl
from openpyxl.chart import BarChart, Reference


def _extract_record_from_sheet(sheet):
    rows = sheet.rows
    next(rows)              # skip header
    res = [
        [c.value for c in row] for row in rows
    ]
    return res


def get_records():
    wb = openpyxl.load_workbook('src_table.xlsx')
    return (_extract_record_from_sheet(wb['Products']),
            _extract_record_from_sheet(wb['Sales']))


def handle_record_process(record: list, year: str):
    wb = openpyxl.load_workbook('src_table.xlsx')
    sales_sheet = wb['Sales']
    record_id, record_name, _ , record_price = record
    wb.create_sheet(record_name + '_sales_info')
    res_sheet = wb[record_name + '_sales_info']
    build_res_sheet(record_id, float(record_price), sales_sheet, res_sheet, year, wb)


def build_res_sheet(prod_id, prod_price, sales_sheet, res_sheet, year: str, wb):
    sales_records = _extract_record_from_sheet(sales_sheet)
    res_map = {i + 1: 0 for i in range(12)}

    for rec in sales_records:
        if rec[0] == prod_id:
            cur_day, cur_month, cur_year = rec[1].split('.')
            prod_amt = int(rec[2])

            if cur_year == year:
                res_map[int(cur_month)] +=  prod_amt * prod_price

    records_to_write = [(month, val) for month, val in res_map.items()]

    res_sheet.append(['20' + year, 'amt'])
    for record in records_to_write:
        res_sheet.append(record)

    chart = BarChart()
    xdata = Reference(res_sheet, min_row=2, min_col=1, max_row=12, max_col=1)
    ydata = Reference(res_sheet, min_row=2, min_col=2, max_row=12, max_col=2)
    chart.set_categories(xdata)
    chart.add_data(ydata)
    res_sheet.add_chart(chart)


    wb.save('src_table_upd.xlsx')


class MainWindow(Tk):

    def __init__(self):
        Tk.__init__(self)
        self._chosen = []
        self._input = StringVar()
        self._make_widgets()

    def _make_widgets(self):
        """ Створити віджети
        """
        self.title('Test program')

        self._nb_frame = ttk.Frame(self)

        tmp = ttk.Entry(self, textvariable=self._input)
        tmp.grid(row=2, column=1, padx=10)
        self._1tree_frame = ttk.Frame(self._nb_frame)
        self._2tree_frame = ttk.Frame(self._nb_frame)
        self._nb_frame.grid(row=1, column=1, columnspan=2,padx=10)
        self._1tree_frame.grid(row=1, column=1, pady=10)
        self._2tree_frame.grid(row=1, column=2, pady=10)

        scroll = ttk.Scrollbar(self._1tree_frame)
        self._costs_tree = ttk.Treeview(self._1tree_frame, height=8, show='headings', yscrollcommand=scroll.set)
        scroll.config(command=self._costs_tree.yview)
        scroll.pack(side=RIGHT, fill=Y)
        self._costs_tree.pack(side=RIGHT)
        self._costs_tree['columns'] = ('id', 'Name', 'Unit', 'Price')
        self._costs_tree.column("id", width=120, anchor='center')
        self._costs_tree.heading("id", text="id", anchor='center')

        self._costs_tree.column("Name", width=120, anchor='center')
        self._costs_tree.heading("Name", text="Name", anchor='center')

        self._costs_tree.column("Unit", width=120, anchor='center')
        self._costs_tree.heading("Unit", text="Unit", anchor='center')

        self._costs_tree.column("Price", width=120, anchor='center')
        self._costs_tree.heading("Price", text="Price", anchor='center')

        self._costs_tree.bind('<Double-1>', self._change_str1)

        scroll = ttk.Scrollbar(self._2tree_frame)
        self._revenue_tree = ttk.Treeview(self._2tree_frame, height=8, show='headings', yscrollcommand=scroll.set)
        scroll.config(command=self._revenue_tree.yview)
        scroll.pack(side=RIGHT, fill=Y)
        self._revenue_tree.pack(side=RIGHT)
        self._revenue_tree['columns'] = ('id', 'Date', 'Quantity')

        self._revenue_tree.column("id", width=120, anchor='center')
        self._revenue_tree.heading("id", text="id", anchor='center')

        self._revenue_tree.column("Date", width=120, anchor='center')
        self._revenue_tree.heading("Date", text="Date", anchor='center')

        self._revenue_tree.column("Quantity", width=120, anchor='center')
        self._revenue_tree.heading("Quantity", text="Quantity", anchor='center')

        self._revenue_tree.bind('<Double-1>', self._change_str2)

        ttk.Button(self, text='Add', command=self._ok_handler).grid(row=2, column=2, padx=10)

        items1, items2 = get_records()
        for item in items1:
            self._costs_tree.insert('', 'end', text='', values=item)

        for item in items2:
            self._revenue_tree.insert('', 'end', text='', values=item)

    def _ok_handler(self, ev=None):
        if self._chosen:
            self._year = self._input.get()
            print(self._chosen, self._year)
            handle_record_process(self._chosen, str(self._year))
            print('done')

    def _change_str1(self, ev=None):
        self._chosen = self._costs_tree.item(self._costs_tree.focus())['values']
        print(self._chosen)

    def _change_str2(self, ev=None):
        self._chosen = self._revenue_tree.item(self._revenue_tree.focus())['values']
        print(self._chosen)


test = MainWindow()
test.mainloop()