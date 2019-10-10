#!/usr/bin/env python3
# -*-encoding: utf-8-*-

from openpyxl import *
from openpyxl.chart import (
    BarChart,
    Reference,
    Series,
)


def plotfunc2():
    wb = load_workbook('test23_22.xlsx')
    ws = wb['Test']  # вибрати активний робочий аркуш

    # створити графік
    chart1 = BarChart()
    chart1.legend = None

    xdata = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    ydata = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    s = Series(ydata, xvalues=xdata)
    chart1.append(s)

    ws.add_chart(chart1, "E1")  # додати графік
    wb.save('graphics.xlsx')  # зберегти робочу книгу


if __name__ == '__main__':
    plotfunc2()
