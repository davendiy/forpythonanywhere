#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl


class IterXLS:

    def __init__(self, doc_name, sheet_name):
        self.doc_name = doc_name
        self.sheet_name = sheet_name
        self._titles = []
        self._row = 2
        self._prepare()

    def __iter__(self):

        _wb = openpyxl.load_workbook(self.doc_name)
        _ws = _wb[self.sheet_name]

        for row in range(2, _ws.max_row):
            tmp = [_ws.cell(row=row, column=i).value
                   for i in range(1, _ws.max_column+1)]
            yield {title: value for title, value in zip(self._titles, tmp)}

        _wb.close()

    def _prepare(self):
        _wb = openpyxl.load_workbook(self.doc_name)
        _ws = _wb[self.sheet_name]

        self._titles = [_ws.cell(row=1, column=i).value
                        for i in range(1, _ws.max_column+1)]
        _wb.close()

    def get_titles(self):
        return self._titles


if __name__ == '__main__':
    for el in IterXLS('test23_22.xlsx', 'Products'):
        print(el)
