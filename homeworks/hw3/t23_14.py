#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from collections import defaultdict

# ----------------------------- sheet names ------------------------------------
PROVIDERS_SHEET = 'Постачальники'
PRODUCTION_SHEET = 'Продукція'
PRICE_SHEET = 'Ціна'

# --------------------------- PROVIDERS SHEET ----------------------------------
PROV_ID_COLUMN = 1
PROV_NAME_COLUMN = 2
PROV_RATING_COLUMN = 3
PROV_ADDRESS_COLUMN = 4

# -------------------------- PRODUCTION SHEET ----------------------------------
PROD_ID_COLUMN = 1
PROD_NAME_COLUMN = 2

# ------------------------------ PRICE SHEET -----------------------------------
S_ID_COLUMN = 1
P_ID_COLUMN = 2
PRICE_COLUMN = 3
TERM_COLUMN = 4


# -------------------mutable variants of named tuples---------------------------
class ProdElem:

    def __init__(self, max_price, all_prices):
        self.max_price = max_price
        self.all_prices = all_prices

    def __repr__(self):
        return f"ProdElem({self.max_price}, {self.all_prices})"

    def __str__(self):
        return repr(self)


class ProvElem:

    def __init__(self, name, rating):
        self.name = name
        self.rating = rating

    def __repr__(self):
        return f"ProvElem({self.name}, {self.rating})"

    def __str__(self):
        return repr(self)


# factory for defaultdict
def _dict_factory():
    return ProdElem(0, {})


def _make_prod_translator(prod_sheet, require_prods):
    """Create necessary data structure from exel file.

    Parameters
    ----------
    prod_sheet    - worksheet from openpyxl
    require_prods - required product names the translator is being made for

    Returns
    -------
    {<prod_id (str)>: <prod_name> (str)}
    """
    translator = {}
    for row in range(2, prod_sheet.max_row+1):
        tmp_name = prod_sheet.cell(row=row, column=PROD_NAME_COLUMN).value
        tmp_id = prod_sheet.cell(row=row, column=PROD_ID_COLUMN).value
        if tmp_name in require_prods:
            translator[tmp_id] = tmp_name

    return translator


def make_main_structure(price_sheet, translator) -> dict:
    """Create data structure from exel file.

    Parameters
    ----------
    price_sheet   - worksheet from openpyxl
    translator    - result of make_prod_translator

    Returns
    -------
    {
        <prod_id (str)>: ProdElem(<max_price (int)>,
                                  {<prov_id (str)>: <price (float)>})
    }
    """
    res_dict = defaultdict(_dict_factory)

    for row in range(2, price_sheet.max_row+1):
        tmp_prod_id = price_sheet.cell(row=row, column=P_ID_COLUMN).value
        tmp_prov_id = price_sheet.cell(row=row, column=S_ID_COLUMN).value
        tmp_price = int(price_sheet.cell(row=row, column=PRICE_COLUMN).value)
        if tmp_prod_id in translator:
            res_dict[tmp_prod_id].max_price = \
                max(res_dict[tmp_prod_id].max_price, tmp_price)
            res_dict[tmp_prod_id].all_prices[tmp_prov_id] = tmp_price
    return res_dict


def _make_prov_dict(prov_sheet):
    """ Creates necessary data structure from exel file.

    Parameters
    ----------
    prov_sheet  - worksheet from openpyxl

    Returns
    -------
    (<max_rating (str)>,
     {<prov_id (str)>: ProvElem(<name (str)>, <rating (int)>)}
    )
    """
    providers = {}
    max_rating = 0
    for row in range(2, prov_sheet.max_row+1):
        tmp_id = prov_sheet.cell(row=row, column=PROV_ID_COLUMN).value
        tmp_name = prov_sheet.cell(row=row, column=PROV_NAME_COLUMN).value
        tmp_rating = int(prov_sheet.cell(row=row, column=PROV_RATING_COLUMN).value)
        providers[tmp_id] = ProvElem(tmp_name, tmp_rating)
        max_rating = max(max_rating, tmp_rating)
    return max_rating, providers


def calculate(filename: str, require_prods: set, a1: float, a2: float):
    """ Main function for finding the best providers of necessary
    production.

    Parameters
    ----------
    filename        - name of exel file with data
    require_prods   - set of required prod names
    a1, a2          - parameters of calculating

    Returns
    -------
    {<required name (str)>: <best provider (str)>}
    """
    wb = openpyxl.load_workbook(filename)
    prov = wb[PROVIDERS_SHEET]
    prod = wb[PRODUCTION_SHEET]
    price = wb[PRICE_SHEET]

    # {<prod_id>: <prod_name>}
    prod_translator = _make_prod_translator(prod, require_prods)

    # {<prod_id>: (<max_price>, {<prov_id>: <price>})}
    production = make_main_structure(price, prod_translator)

    # {prov_id: (<name>, <rating>)}
    max_rating, providers = _make_prov_dict(prov)
    res_dict = {}
    for prod_id, prod_inf in production.items():
        best_prov_id = ''
        best_prov = 0
        for prov_id, price in prod_inf.all_prices.items():
            tmp_prov = a1 * price / prod_inf.max_price \
                       + a2 * providers[prov_id].rating / max_rating
            if tmp_prov > best_prov:
                best_prov = tmp_prov
                best_prov_id = prov_id
        res_dict[prod_translator[prod_id]] = providers[best_prov_id].name
    return res_dict


if __name__ == '__main__':
    print(calculate('test23_14.xlsx', {"Олівець", "Ручка"}, 12, 14))
