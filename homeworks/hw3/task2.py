#!/usr/bin/env python3
# -*-encoding: utf-8-*-

# created: 10.10.2019
# by David Zashkolny
# 3 course, comp math
# Taras Shevchenko National University of Kyiv
# email: davendiy@gmail.com

import openpyxl
from string import ascii_uppercase
import re


RE_RANGE = re.compile(r"=SUM\((?P<start_y>[A-Z]+)(?P<start_x>[0-9]+):"
                      r"(?P<end_y>[A-Z]+)(?P<end_x>[0-9]+)\)")


def cells(exel_range: str):
    start_y, start_x, end_y, end_x = RE_RANGE.match(exel_range).groups()

    if not (start_x and start_y and end_x and end_y):
        return

    start_y = 1 + ascii_uppercase.index(start_y[0]) + \
              (len(start_y)-1) * len(ascii_uppercase)
    end_y = 1 + ascii_uppercase.index(end_y[0]) + \
            (len(end_y)-1) * len(ascii_uppercase)
    for x in range(int(start_x), int(end_x)+1):
        for y in range(start_y, end_y+1):
            yield (x, y)


def calc(worksheet, range):
    res = 0
    for row, column in cells(range):
        try:
            res += float(worksheet.cell(row=row, column=column).value)
        except Exception as e:
            print(e)
            continue
    return res


def exel_handler(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    res_ws = wb.create_sheet(sheet_name + '_numerical_only')
    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            tmp_data = ws.cell(row=row, column=col).value   # type: str

            if isinstance(tmp_data, int):
                res_ws.cell(row=row, column=col).value = tmp_data
            elif isinstance(tmp_data, str) and tmp_data.startswith("=SUM"):
                res_ws.cell(row=row, column=col).value = calc(ws, tmp_data)
    wb.save(filename)


if __name__ == '__main__':
    exel_handler('test_task2.xlsx', 'test')
